
import openpyxl

def getrowCount(file,sheetname):
    book = openpyxl.load_workbook(file)
    sheet = book[sheetname]
    return sheet.max_row

def getreadData(file,sheetname,row_no,col_no):
    book = openpyxl.load_workbook(file)
    sheet = book[sheetname]
    return sheet.cell(row=row_no, column=col_no).value

def getwriteData(file,sheetname,row_no,col_no,data):
    book = openpyxl.load_workbook(file)
    sheet = book[sheetname]
    sheet.cell(row=row_no,column=col_no).value = data
    book.save(file)
    




